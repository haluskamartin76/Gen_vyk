import streamlit as st
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import date
import calendar
import io

# --- KONFIGURÁCIA ---
MESIACE_SK = ["Január", "Február", "Marec", "Apríl", "Máj", "Jún", "Júl", "August", "September", "Október", "November", "December"]
DNI_SK = ["Pondelok", "Utorok", "Streda", "Štvrtok", "Piatok", "Sobota", "Nedeľa"]

st.set_page_config(page_title="Generátor Výkazu NR SR", layout="wide")

# Funkcia na výpočet sviatkov (vrátane Veľkej noci)
def ziskaj_sviatky(rok, mesiac):
    sviatky = {1: [1, 6], 5: [1, 8], 7: [5], 8: [29], 9: [1, 15], 11: [1, 17], 12: [24, 25, 26]}
    velka_noc = {
        2024: [(3, 29), (4, 1)], 2025: [(4, 18), (4, 21)], 2026: [(4, 3), (4, 6)],
        2027: [(3, 26), (3, 29)], 2028: [(4, 14), (4, 17)], 2029: [(3, 30), (4, 2)], 2030: [(4, 19), (4, 22)]
    }
    mes_sviatky = sviatky.get(mesiac, []).copy()
    if rok in velka_noc:
        for m, d in velka_noc[rok]:
            if m == mesiac: mes_sviatky.append(d)
    return mes_sviatky

if 'reset_counter' not in st.session_state:
    st.session_state.reset_counter = 0

def reset_cycle():
    st.session_state.reset_counter += 1

st.title("📄 Generátor pracovného výkazu")

# --- BOČNÝ PANEL ---
with st.sidebar:
    st.header("Nastavenia")
    meno = st.text_input("Meno a Priezvisko:", placeholder="Martin Haluska")
    mesiac_meno = st.selectbox("Mesiac:", MESIACE_SK, index=date.today().month - 1)
    mesiac_idx = MESIACE_SK.index(mesiac_meno) + 1
    rok_val = st.selectbox("Rok:", list(range(2024, 2031)), index=2)
    zmena_skupina = st.selectbox("Základný cyklus:", ["Zmena 1", "Zmena 2", "Zmena 3", "Zmena 4"], index=0)
    fond = st.number_input("Fond hodín na mesiac:", value=154.0, step=0.5)
    
    st.divider()
    col_edit1, col_edit2 = st.columns([0.8, 0.2])
    with col_edit2:
        edit_mode = st.checkbox("✏️", help="Kliknite pre úpravu útvaru")
    with col_edit1:
        utvar_val = st.text_input("Organizačný útvar:", 
                             value="Odbor obrany, bezpečnosti a ochrany", 
                             disabled=not edit_mode)
    
    st.divider()
    st.button("🔄 Načítať / Resetovať cyklus", use_container_width=True, on_click=reset_cycle)

# --- LOGIKA CYKLU ---
cykly = {"Zmena 1": "DNVDNVVV", "Zmena 2": "VVDNVDNV", "Zmena 3": "VDNVVVDN", "Zmena 4": "NVVVDNVD"}
vzor, ref = cykly[zmena_skupina], date(2026, 3, 1)
_, dni_count = calendar.monthrange(rok_val, mesiac_idx)
aktualne_sviatky = ziskaj_sviatky(rok_val, mesiac_idx)

st.subheader(f"Upresnenie dní pre {mesiac_meno} {rok_val}")

# --- INTERAKTÍVNA TABUĽKA ---
vysledne_dni = []
h1, h2, h3, h4, h5 = st.columns([0.6, 0.6, 0.6, 1, 2.5])
with h1: st.markdown("**Deň**")
with h2: st.markdown("**D**")
with h3: st.markdown("**N**")
with h4: st.markdown("**Absencia**")
with h5: st.markdown("**Dátum**")
st.markdown("---")

for d in range(1, dni_count + 1):
    dt = date(rok_val, mesiac_idx, d)
    pos = (dt - ref).days % 8
    is_weekend = dt.weekday() >= 5
    is_sviatok = d in aktualne_sviatky
    
    with st.container():
        col1, col2, col3, col4, col5 = st.columns([0.6, 0.6, 0.6, 1, 2.5])
        with col1: st.write(f"{d}.")
        with col2: d_val = st.checkbox(" ", value=(vzor[pos] == "D"), key=f"d_{d}_{st.session_state.reset_counter}", label_visibility="collapsed")
        with col3: n_val = st.checkbox(" ", value=(vzor[pos] == "N"), key=f"n_{d}_{st.session_state.reset_counter}", label_visibility="collapsed")
        with col4: abs_val = st.text_input("Abs", key=f"a_{d}_{st.session_state.reset_counter}", label_visibility="collapsed", placeholder="D/KZ/PN")
        with col5: 
            den_text = f"{dt.strftime('%d.%m.')} ({DNI_SK[dt.weekday()]})"
            if is_sviatok: st.markdown(f"<span style='color:#0066cc; font-weight:bold;'>{den_text} 🚩</span>", unsafe_allow_html=True)
            elif is_weekend: st.markdown(f"<span style='color:orange; font-weight:bold;'>{den_text}</span>", unsafe_allow_html=True)
            else: st.write(den_text)
        
        vysledne_dni.append({"den": d, "is_d": d_val, "is_n": n_val, "abs": abs_val.upper().strip(), "weekday": dt.weekday()})
        st.markdown("<hr style='margin:0; border-top: 1px solid #eee;'>", unsafe_allow_html=True)

# --- GENEROVANIE EXCELU ---
st.write("")
if st.button("💾 VYGENEROVAŤ EXCEL VÝKAZ", type="primary", use_container_width=True):
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Definícia farieb optimalizovaných pre ČB tlač
    fill_sat = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid") # Svetlošedá (Sobota)
    fill_sun = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid") # Stredne šedá (Nedeľa)
    fill_hol = PatternFill(start_color="595959", end_color="595959", fill_type="solid") # Tmavosivá (Sviatok)

    f_header = Font(name='Arial', size=11, bold=True)
    f_bold = Font(name='Arial', size=9, bold=True)
    f_norm = Font(name='Arial', size=9)
    f_white = Font(name='Arial', size=9, color="FFFFFF", bold=True) # Biele písmo pre kontrast na tmavom pozadí
    
    align_c = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Hlavička dokumentu
    ws.merge_cells('A1:J1'); ws['A1'] = "Kancelária Národnej rady Slovenskej republiky"; ws['A1'].font = f_header; ws['A1'].alignment = align_c
    ws.merge_cells('A2:J2'); ws['A2'] = f"Pracovný výkaz za mesiac {mesiac_meno} {rok_val}"; ws['A2'].font = f_header; ws['A2'].alignment = align_c
    ws['A4'], ws['C4'], ws['H4'], ws['I4'] = "Organizačný útvar:", utvar_val, "Fond:", fond
    ws['A6'], ws['C6'] = "Zamestnanec:", meno
    for c in ['C4', 'I4', 'C6']: ws[c].font = f_bold

    # Hlavička tabuľky
    ws.merge_cells('A8:A10'); ws['A8'] = "Dátum"
    ws.merge_cells('B8:B10'); ws['B8'] = "Pracovná doba"
    ws.merge_cells('C8:C10'); ws['C8'] = "Odprac.\nhod.\nspolu"
    ws.merge_cells('D8:H8'); ws['D8'] = "z toho"
    ws.merge_cells('I8:J8'); ws['I8'] = "Pracovná pohotovosť"
    ws.merge_cells('D9:D10'); ws['D9'] = "v so-ne"; ws.merge_cells('E9:E10'); ws['E9'] = "v noci"
    ws.merge_cells('F9:G9'); ws['F9'] = "nadčas"; ws.merge_cells('H9:H10'); ws['H9'] = "sviatok"
    ws['I9'], ws['J9'] = "na\npracovisk.", "mimo\npracovisk."
    ws['F10'], ws['G10'] = "Po – Pia", "So – Ne"

    for r_idx in range(8, 11):
        for c_idx in range(1, 11):
            cell = ws.cell(r_idx, c_idx); cell.font = f_bold; cell.alignment = align_c; cell.border = border

    for row_data in vysledne_dni:
        d, dow = row_data["den"], row_data["weekday"]
        is_sv = d in aktualne_sviatky
        p_doba, h_sp, h_sn, h_no, h_sv = "", "", "", "", ""
        
        if row_data["abs"] == "D": p_doba, h_sp = "Dovolenka", 11.5
        elif row_data["abs"] == "KZ": p_doba, h_sp = "KZ", 11.5
        elif row_data["abs"] in ["PN", "L"]: p_doba = row_data["abs"]
        elif row_data["is_d"]:
            p_doba, h_sp = "06:00 - 18:00", 11.5
            if dow >= 5: h_sn = 11.5
            if is_sv: h_sv = 11.5
        elif row_data["is_n"]:
            p_doba, h_sp, h_no = "18:00 - 06:00", 11.5, 7.5
            if dow == 4: h_sn = 6.0
            elif dow == 5: h_sn = 11.5
            elif dow == 6: h_sn = 5.5
            if is_sv: h_sv = 11.5

        line = [d, p_doba, h_sp, h_sn, h_no, "", "", h_sv, "", ""]
        for idx, val in enumerate(line, 1):
            cell = ws.cell(11+d-1, idx, val); cell.border = border; cell.alignment = align_c; cell.font = f_norm
            if idx == 1:
                if is_sv:
                    cell.fill = fill_hol
                    cell.font = f_white # Biele písmo pre sviatok
                elif dow == 5: cell.fill = fill_sat
                elif dow == 6: cell.fill = fill_sun

    # Päta a vzorce
    s_row = 11 + dni_count
    ws.merge_cells(f'A{s_row}:B{s_row}'); ws[f'A{s_row}'] = "Odpracované hodiny s p o l u:"
    ws[f'A{s_row}'].alignment = Alignment(horizontal='left'); ws[f'A{s_row}'].font = f_bold
    for col_l in ["C", "D", "E", "H", "I", "J"]:
        ws[f'{col_l}{s_row}'] = f"=SUM({col_l}11:{col_l}{s_row-1})"
        ws[f'{col_l}{s_row}'].font = f_bold; ws[f'{col_l}{s_row}'].border = border; ws[f'{col_l}{s_row}'].alignment = align_c
    ws.merge_cells(f'F{s_row}:G{s_row}'); ws[f'F{s_row}'] = f"=C{s_row}-I4"; ws[f'F{s_row}'].font = f_bold; ws[f'F{s_row}'].alignment = align_c
    for c_idx in range(1, 11): ws.cell(s_row, c_idx).border = border
    
    # ZLÚČENÉ PODPISY (OPRAVA VIZUÁLU)
    sig_r = s_row + 3
    ws.merge_cells(f'B{sig_r}:D{sig_r}'); ws[f'B{sig_r}'] = "____________________________"
    ws.merge_cells(f'G{sig_r}:I{sig_r}'); ws[f'G{sig_r}'] = "____________________________"
    ws[f'B{sig_r}'].alignment = align_c; ws[f'G{sig_r}'].alignment = align_c
    
    ws.merge_cells(f'B{sig_r+1}:D{sig_r+1}'); ws[f'B{sig_r+1}'] = "podpis zamestnanca"
    ws.merge_cells(f'G{sig_r+1}:I{sig_r+1}'); ws[f'G{sig_r+1}'] = "podpis vedúceho"
    ws[f'B{sig_r+1}'].alignment = align_c; ws[f'G{sig_r+1}'].alignment = align_c

    widths = [6, 18, 10, 10, 10, 10, 10, 10, 10, 10]
    for i, w in enumerate(widths, 1): ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    wb.save(output)
    st.download_button(label="📥 STIAHNUŤ EXCEL", data=output.getvalue(), file_name=f"Vykaz_{meno.replace(' ','_')}_{mesiac_meno}_{rok_val}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
